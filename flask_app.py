#http://localhost:5000/increase?uplift=10  #sample URI
#import relevant libraries
from flask import Flask, request
import openpyxl

#the application object
app = Flask(__name__)

#home 
@app.route('/')
def hello_world():
	return 'Welcome to Sarahs Week 5 Fruit Shop Assignment!'

#price increase API
@app.route('/increase')
def increase(): 
	args=request.args       #reading the arguments at the end of URL
	variables = dict(args)  #moving arguments from strings to variables

	uplift=float(variables['uplift'])              #converting variable to float (decimals)
	wb = openpyxl.load_workbook("Pricelist.xlsx")  #locating the relevant excel sheet
	ws = wb.active								   #selecting the active sheet

	npl={}   									   #creating a json object (to store new price list) cause its easy to handle :)
	max_rows = ws.max_row  						   #no matter how many fruits you add, it will calculate the price for all of them!

	for row in ws.iter_rows(min_row=2, max_row=max_rows, min_col=1, max_col=2):    #iterating through products and prices
		for cell in row:														   #processing each cell in every row
			if isinstance((cell.value), (str)):									   #if the cell we are processing is a string:
				print(cell.value)												   #print the cell
				key = cell.value											   	   #creating the word for the dictionary
			else:															 	   #otherwise,
				floatyPrice=float(cell.value)								 	   #convert the price to a floating decimal
				upliftPrice=round(floatyPrice*uplift,2)   					 	   #multiply original price with variable uplift (defined in postman in this case), and rounded to 2 dp
				print(upliftPrice) 											       #print the updated price (for testing script)
				npl[key] = upliftPrice										       #adding the definition to the word for this iteration

	return dict(npl) 														       #the response! can be seen on postman (or browser: http://localhost:5000/increase?uplift=10)

def main():
	app.run()

if __name__== '__main__':
	main()


